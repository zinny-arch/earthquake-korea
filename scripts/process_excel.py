"""
지진 데이터를 읽어 가독성 높은 Excel 파일로 출력합니다.
"""
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

DATA_FILE = os.path.join(os.path.dirname(__file__), "..", "data", "raw", "earthquake_5years.csv")
OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "..", "output", "지진데이터_2021-2026.xlsx")

# ── 색상 팔레트 ──────────────────────────────────────────────────────────────
HEADER_BG    = "1F4E79"
HEADER_FG    = "FFFFFF"
SUBHDR_BG    = "2E75B6"
TITLE_BG     = "0D3349"
ROW_ALT      = "EBF3FB"
ROW_NORMAL   = "FFFFFF"
BORDER_COLOR = "BDD7EE"

MAG_LOW    = "C6EFCE"  # < 2.0  (연두)
MAG_MID    = "FFEB9C"  # 2.0~3.0 (노랑)
MAG_HIGH   = "FFC7CE"  # ≥ 3.0  (빨강)

THIN = Side(style="thin", color=BORDER_COLOR)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def magnitude_fill(mag: float) -> PatternFill:
    if mag < 2.0:
        color = MAG_LOW
    elif mag < 3.0:
        color = MAG_MID
    else:
        color = MAG_HIGH
    return PatternFill(fill_type="solid", fgColor=color)


def set_header(ws, row, col, value, bg=HEADER_BG, fg=HEADER_FG, bold=True, center=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="맑은 고딕", bold=bold, color=fg, size=10)
    cell.fill = PatternFill(fill_type="solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center" if center else "left",
                               vertical="center", wrap_text=True)
    cell.border = BORDER
    return cell


def write_main_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.active
    ws.title = "지진 목록"
    ws.sheet_view.showGridLines = False

    # ── 제목 행 ───────────────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "대한민국 최근 5년간 지진 목록 (2021–2026)"
    title_cell.font = Font(name="맑은 고딕", bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill(fill_type="solid", fgColor=TITLE_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── 열 헤더 ──────────────────────────────────────────────────────────────
    headers = ["No.", "발생시각", "규모", "위도", "경도", "깊이(km)", "위치", "규모 등급"]
    for col, h in enumerate(headers, 1):
        set_header(ws, 2, col, h)
    ws.row_dimensions[2].height = 22

    # ── 열 너비 ──────────────────────────────────────────────────────────────
    col_widths = [6, 22, 9, 10, 11, 10, 22, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── 데이터 행 ────────────────────────────────────────────────────────────
    df = df.copy().reset_index(drop=True)
    if "발생시각" in df.columns:
        df["발생시각"] = pd.to_datetime(df["발생시각"], errors="coerce")

    def mag_label(m):
        try:
            m = float(m)
        except (ValueError, TypeError):
            return "-"
        if m < 2.0:
            return "미소지진"
        elif m < 3.0:
            return "소규모"
        elif m < 4.0:
            return "중규모"
        else:
            return "강진"

    for idx, row_data in df.iterrows():
        r = idx + 3
        bg = ROW_ALT if idx % 2 == 0 else ROW_NORMAL
        base_fill = PatternFill(fill_type="solid", fgColor=bg)

        mag_val = row_data.get("규모", None)
        try:
            mag_float = float(mag_val)
        except (ValueError, TypeError):
            mag_float = None

        values = [
            idx + 1,
            row_data.get("발생시각", ""),
            mag_float,
            row_data.get("위도", ""),
            row_data.get("경도", ""),
            row_data.get("깊이(km)", ""),
            row_data.get("위치", ""),
            mag_label(mag_float) if mag_float else "-",
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = Font(name="맑은 고딕", size=9)
            cell.border = BORDER
            cell.alignment = Alignment(
                horizontal="center" if col != 7 else "left",
                vertical="center"
            )
            # 규모 셀만 색상 구분
            if col == 3 and mag_float is not None:
                cell.fill = magnitude_fill(mag_float)
            elif col == 8 and mag_float is not None:
                cell.fill = magnitude_fill(mag_float)
            else:
                cell.fill = base_fill

            # 날짜 형식
            if col == 2 and isinstance(val, pd.Timestamp):
                cell.number_format = "YYYY-MM-DD HH:MM:SS"

            # 소수 형식
            if col in (3, 4, 5):
                cell.number_format = "0.0##"

        ws.row_dimensions[r].height = 16

    # 틀 고정 (헤더 고정)
    ws.freeze_panes = "A3"


def write_stats_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("지역별 통계")
    ws.sheet_view.showGridLines = False

    if "위치" not in df.columns or "규모" not in df.columns:
        ws["A1"] = "데이터 없음"
        return

    df["규모"] = pd.to_numeric(df["규모"], errors="coerce")
    stats = (
        df.groupby("위치")["규모"]
        .agg(발생횟수="count", 평균규모="mean", 최대규모="max", 최소규모="min")
        .sort_values("발생횟수", ascending=False)
        .reset_index()
    )
    stats["평균규모"] = stats["평균규모"].round(2)

    ws.merge_cells("A1:F1")
    title = ws["A1"]
    title.value = "지역별 지진 통계"
    title.font = Font(name="맑은 고딕", bold=True, size=13, color="FFFFFF")
    title.fill = PatternFill(fill_type="solid", fgColor=TITLE_BG)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    stat_headers = ["지역", "발생횟수", "평균규모", "최대규모", "최소규모", "비율(%)"]
    for col, h in enumerate(stat_headers, 1):
        set_header(ws, 2, col, h)

    total = stats["발생횟수"].sum()
    for idx, row_data in stats.iterrows():
        r = idx + 3
        bg = ROW_ALT if idx % 2 == 0 else ROW_NORMAL
        fill = PatternFill(fill_type="solid", fgColor=bg)
        ratio = round(row_data["발생횟수"] / total * 100, 1)
        values = [
            row_data["위치"],
            row_data["발생횟수"],
            row_data["평균규모"],
            row_data["최대규모"],
            row_data["최소규모"],
            ratio,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = Font(name="맑은 고딕", size=9)
            cell.border = BORDER
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center" if col != 1 else "left",
                                       vertical="center")
            if col in (3, 4, 5):
                cell.number_format = "0.0"
            if col == 6:
                cell.number_format = "0.0\"%\""
        ws.row_dimensions[r].height = 16

    col_widths_s = [18, 12, 12, 12, 12, 12]
    for i, w in enumerate(col_widths_s, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"


def write_yearly_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("연도별 통계")
    ws.sheet_view.showGridLines = False

    df = df.copy()
    df["규모"] = pd.to_numeric(df["규모"], errors="coerce")
    df["발생시각"] = pd.to_datetime(df["발생시각"], errors="coerce")
    df["연도"] = df["발생시각"].dt.year

    yearly = (
        df.groupby("연도")["규모"]
        .agg(발생횟수="count", 평균규모="mean", 최대규모="max")
        .reset_index()
    )
    yearly["평균규모"] = yearly["평균규모"].round(2)

    ws.merge_cells("A1:E1")
    title = ws["A1"]
    title.value = "연도별 지진 통계"
    title.font = Font(name="맑은 고딕", bold=True, size=13, color="FFFFFF")
    title.fill = PatternFill(fill_type="solid", fgColor=TITLE_BG)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for col, h in enumerate(["연도", "발생횟수", "평균규모", "최대규모", "전년대비"], 1):
        set_header(ws, 2, col, h)

    prev_count = None
    for idx, row_data in yearly.iterrows():
        r = idx + 3
        count = row_data["발생횟수"]
        if prev_count is not None:
            diff = count - prev_count
            diff_str = f"+{diff}" if diff >= 0 else str(diff)
        else:
            diff_str = "-"
        prev_count = count

        bg = ROW_ALT if idx % 2 == 0 else ROW_NORMAL
        fill = PatternFill(fill_type="solid", fgColor=bg)
        for col, val in enumerate([row_data["연도"], count, row_data["평균규모"],
                                    row_data["최대규모"], diff_str], 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = Font(name="맑은 고딕", size=9)
            cell.border = BORDER
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col == 3:
                cell.number_format = "0.0"
        ws.row_dimensions[r].height = 16

    for i, w in enumerate([10, 12, 12, 12, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"


def main():
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)

    print(f"데이터 로드: {DATA_FILE}")
    df = pd.read_csv(DATA_FILE, encoding="utf-8-sig")
    print(f"  → {len(df)}건")

    wb = Workbook()
    write_main_sheet(wb, df)
    write_stats_sheet(wb, df)
    write_yearly_sheet(wb, df)

    wb.save(OUTPUT_FILE)
    print(f"Excel 저장 완료: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
